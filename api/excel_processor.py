import json
import datetime
import re as regex
from openpyxl import load_workbook
from ftfy import fix_text
from .utils import DateTimeEncoder

def clean_text(text):
    """Limpa e corrige a codificação de uma string, preservando objetos de data."""
    # Se for uma data, retorna o objeto diretamente para ser formatado depois
    if isinstance(text, datetime.datetime):
        return text

    value = str(text if text is not None else '')
    try:
        cleaned_value = fix_text(value)
        return ' '.join(cleaned_value.split())
    except Exception:
        return ' '.join(value.split())

def process_single_excel_to_json(file_path):
    """Processa um único ficheiro Excel e retorna os dados e o índice do cabeçalho."""
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows():
            data.append([cell.value for cell in row])

        header_row_index = -1
        possible_headers_legal_one = {h.lower() for h in [
            'Pasta', 'Tipo', 'Data de distribuição', 'Número antigo', 'Número de CNJ', 'Ação',
            'Cliente principal / Contato / Nome/Razão social', 'Cliente principal / Posição',
            'Cidade', 'UF', 'Status', 'Andamentos / Data/hora', 'Andamentos / Descrição'
        ]}
        possible_headers_advise = {h.lower() for h in [
            'DIÁRIO', 'PROCESSO', 'PUBLICAÇÃO EM', 'COMARCA', 'VARA', 'DISPONIBILIZAÇÃO EM',
            'PALAVRA CHAVE', 'CADERNO', 'CONTRATANTE', 'USUÁRIO', 'EDIÇÃO', 'PÁGINA INICIAL',
            'PÁGINA FINAL', 'DESPACHO', 'CONTEUDO'
        ]}

        for i, row in enumerate(data[:5]):
            if not any(cell for cell in row):
                continue
            row_values = {str(cell).strip().lower() for cell in row if cell}
            score_legal_one = len(row_values.intersection(possible_headers_legal_one))
            score_advise = len(row_values.intersection(possible_headers_advise))
            if score_legal_one > 5 or score_advise > 5:
                header_row_index = i
                break

        if header_row_index == -1:
            raise ValueError(f"Não foi possível encontrar um cabeçalho válido em {file_path}")
        
        header = [clean_text(str(cell)) for cell in data[header_row_index]]
        json_data = []
        for pub_number, row in enumerate(data[header_row_index + 1:], 1):
            if not any(cell is not None for cell in row):
                continue
            row_data = {}
            row_data["numero da publicação"] = pub_number
            for col_name, cell_value in zip(header, row):
                if 'unnamed' in col_name.lower() or not col_name:
                    continue
                row_data[col_name] = clean_text(cell_value)
            if 'None' in row_data:
                del row_data['None']
            if len(row_data) > 1:
                json_data.append(row_data)
        return json_data, header_row_index

    except FileNotFoundError as e:
        print(f"Erro: Arquivo não encontrado - {e}")
        return [], -1
    except Exception as e:
        print(f"Ocorreu um erro inesperado ao processar o arquivo {file_path}: {e}")
        return [], -1

def process_excel_files_from_paths(file_paths: list):
    """Função principal que orquestra o processamento de uma lista de ficheiros Excel."""
    unified_data = []

    for file in file_paths:
        processed_data, header_index = process_single_excel_to_json(file)
        if not processed_data:
            continue
        
        origem = "Adivise" if header_index == 0 else "Legal One"
        
        for item in processed_data:
            item['origem'] = origem
        
        unified_data.extend(processed_data)

    # Listas de chaves a serem removidas, agora incluindo as novas solicitações
    keys_to_remove_legal_one = ['Pasta', 'Tipo', 'Número antigo', 'Cliente principal / Contato / Nome/Razão social', 'Cliente principal / Posição', 'Cidade', 'UF', 'Status', 'Andamentos / Data/hora']
    keys_to_remove_advise = ['DIÁRIO', 'COMARCA', 'VARA', 'PALAVRA CHAVE', 'CADERNO', 'CONTRATANTE', 'USUÁRIO', 'EDIÇÃO', 'PÁGINA INICIAL', 'PÁGINA FINAL', 'DISPONIBILIZAÇÃO EM']

    # Processamento inicial e enriquecimento dos dados
    for item in unified_data:
        if item.get('origem') == 'Legal One':
            if not item.get('Número de CNJ'):
                descricao = item.get('Andamentos / Descrição', '')
                match = regex.search(r'N[ÚU]MERO [ÚU]NICO:?\s*([\d.-]+)', descricao, regex.IGNORECASE)
                if match:
                    item['Número de CNJ'] = match.group(1)
                else:
                    item['Número de CNJ'] = 'Não encontrado'
            
            if item.get('Número de CNJ') and item['Número de CNJ'] != 'Não encontrado':
                item['Número de CNJ'] = regex.sub(r'[.\-]', '', item['Número de CNJ'])

            for key in keys_to_remove_legal_one:
                item.pop(key, None)

        elif item.get('origem') == 'Adivise':
            if item.get('PROCESSO'):
                item['PROCESSO'] = regex.sub(r'[.\-]', '', item['PROCESSO'])

            for key in keys_to_remove_advise:
                item.pop(key, None)

    # Etapa final de padronização dos nomes das chaves
    standardized_data = []
    for item in unified_data:
        standard_item = {}
        standard_item['ID'] = item.get('numero da publicação')
        
        # Unifica 'Número de CNJ' e 'PROCESSO' em 'Processo'
        if item.get('origem') == 'Legal One':
            standard_item['Processo'] = item.get('Número de CNJ')
            date_value = item.get('Data de distribuição')
            standard_item['Ação/ Despacho'] = item.get('Ação')
            standard_item['Conteudo'] = item.get('Andamentos / Descrição')
        elif item.get('origem') == 'Adivise':
            standard_item['Processo'] = item.get('PROCESSO')
            date_value = item.get('PUBLICAÇÃO EM')
            standard_item['Ação/ Despacho'] = item.get('DESPACHO')
            standard_item['Conteudo'] = item.get('CONTEUDO')

        # Formata a data para DD/MM/YYYY, se for um objeto datetime
        if isinstance(date_value, datetime.datetime):
            standard_item['Data Publicação'] = date_value.strftime('%d/%m/%Y')
        else:
            standard_item['Data Publicação'] = date_value # Mantém como está se não for uma data

        standard_item['Origem'] = item.get('origem')
        standardized_data.append(standard_item)

    return standardized_data
