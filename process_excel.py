"""
Lê um arquivo Excel usando openpyxl, converte cada linha em um objeto JSON e retorna uma lista de objetos.
"""
import json
import datetime
import re as regex
from openpyxl import load_workbook
from ftfy import fix_text
from utils import DateTimeEncoder

def clean_text(text):
    """Limpa e corrige a codificação de uma string de forma robusta."""
    value = str(text if text is not None else '')
    try:
        # Tenta corrigir a codificação e outros problemas comuns
        cleaned_value = fix_text(value)
        # Remove espaços extras
        return ' '.join(cleaned_value.split())
    except Exception:
        # Se ftfy falhar, retorna o valor original apenas com espaços limpos.
        return ' '.join(value.split())

def process_excel_to_json(file_path):
    """
    Lê um arquivo Excel usando openpyxl, converte cada linha em um objeto JSON e retorna uma lista de objetos.
    """
    try:
        workbook = load_workbook(file_path) # CORREÇÃO: Chamada direta da função
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows():
            data.append([cell.value for cell in row])

        # --- Lógica de Detecção Automática de Cabeçalho ---
        header_row_index = -1
        possible_headers_legal_one = {h.lower() for h in [
            'Pasta', 'Tipo', 'Data de distribuição', 'Número antigo', 'Número de CNJ', 'Ação',
            'Cliente principal / Contato / Nome/Razão social', 'Cliente principal / Posição',
            'Cidade', 'UF', 'Status', 'Andamentos / Data/hora', 'Andamentos / Descrição'
        ]}
        possible_headers_advise = {h.lower() for h in [
            'DIÁRIO', 'PROCESSO', 'PUBLICAÇÃO EM', 'COMARCA', 'VARA', 'DISPONIBILIZAÇÃO EM',
            'PALAVRA CHAVE', 'CADERNO', 'CONTRATANTE', 'USUÁRIO', 'EDIÇÃO', 'PÁGINA INICIAL',
            'PÁGINA FINAL', 'DESPACHO', 'CONTEUDO'
        ]}

        for i, row in enumerate(data[:5]): # Analisa as 5 primeiras linhas
            if not any(cell for cell in row):
                continue

            row_values = {str(cell).strip().lower() for cell in row if cell}

            score_legal_one = len(row_values.intersection(possible_headers_legal_one))
            score_advise = len(row_values.intersection(possible_headers_advise))

            if score_legal_one > 5 or score_advise > 5:
                header_row_index = i
                break

        if header_row_index == -1:
            raise ValueError(f"Não foi possível encontrar um cabeçalho válido em {file_path}")
        
        # --- Fim da Lógica de Detecção ---
        header = [clean_text(str(cell)) for cell in data[header_row_index]]
        json_data = []
        for pub_number, row in enumerate(data[header_row_index + 1:], 1):
            if not any(cell is not None for cell in row):
                continue
            
            row_data = {}
            row_data["numero da publicação"] = pub_number

            for col_name, cell_value in zip(header, row):
                if 'unnamed' in col_name.lower() or not col_name:
                    continue
                row_data[col_name] = clean_text(cell_value)
            
            if 'None' in row_data:
                del row_data['None']

            if len(row_data) > 1:
                json_data.append(row_data)
        return json_data, header_row_index

    except FileNotFoundError as e:
        print(f"Erro: Arquivo não encontrado - {e}")
        return [], -1
    except Exception as e:
        print(f"Ocorreu um erro inesperado ao processar o arquivo {file_path}: {e}")
        return [], -1

if __name__ == "__main__":
    files_to_process = [
        r'c:\Users\Paulo Martins\Desktop\Candido ADV - Projeto prompt GPT BUILDER\24.09.2025\LEGAL ONE -Relatório Publicações.xlsx',
        r'c:\Users\Paulo Martins\Desktop\Candido ADV - Projeto prompt GPT BUILDER\24.09.2025\ADIVISE -candido-dortas-sociedade-de-advogadas-e-advogados-638943060441979375..xlsx'
    ]
    output_file = 'output.json'
    unified_data = []

    for file in files_to_process:
        print(f"Processando arquivo: {file.split('\\')[-1]}...")
        processed_data, header_index = process_excel_to_json(file)

        # Pula para o próximo arquivo se não houver dados
        if not processed_data:
            continue
        
        # Define a origem com base na linha do cabeçalho detectada
        origem = "Adivise" if header_index == 0 else "Legal One"
        
        # Adiciona a origem a cada registro
        for item in processed_data:
            item['origem'] = origem
        
        unified_data.extend(processed_data)

    # Define as chaves a serem removidas para cada origem
    keys_to_remove_legal_one = ['Pasta', 'Tipo', 'Número antigo', 'Cliente principal / Contato / Nome/Razão social', 'Cliente principal / Posição', 'Cidade', 'UF', 'Status']
    keys_to_remove_advise = ['DIÁRIO', 'COMARCA', 'VARA', 'PALAVRA CHAVE', 'CADERNO', 'CONTRATANTE', 'USUÁRIO', 'EDIÇÃO', 'PÁGINA INICIAL', 'PÁGINA FINAL']

    # Itera sobre a lista unificada para limpar e enriquecer os dados
    for item in unified_data:
        if item.get('origem') == 'Legal One':
            # 1. Preenche 'Número de CNJ' se estiver vazio
            if not item.get('Número de CNJ'):
                descricao = item.get('Andamentos / Descrição', '')
                # Procura por 'Número único' (case-insensitive) e captura o número que vem depois
                match = regex.search(r'N[ÚU]MERO [ÚU]NICO:?\s*([\d.-]+)', descricao, regex.IGNORECASE)
                if match:
                    item['Número de CNJ'] = match.group(1)
                else:
                    item['Número de CNJ'] = 'Não encontrado'
            
            # 2. Remove a pontuação de 'Número de CNJ', se houver
            if item.get('Número de CNJ') and item['Número de CNJ'] != 'Não encontrado':
                item['Número de CNJ'] = regex.sub(r'[.\-]', '', item['Número de CNJ'])

            # 3. Remove as chaves indesejadas do Legal One
            for key in keys_to_remove_legal_one:
                item.pop(key, None)

        elif item.get('origem') == 'Adivise':
            # 1. Remove a pontuação do 'PROCESSO'
            if item.get('PROCESSO'):
                item['PROCESSO'] = regex.sub(r'[.\-]', '', item['PROCESSO'])

            # 2. Remove as chaves indesejadas do Adivise
            for key in keys_to_remove_advise:
                item.pop(key, None)

    # Converte a lista final para JSON e salva no arquivo
    final_json_output = json.dumps(unified_data, indent=4, ensure_ascii=False, cls=DateTimeEncoder)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(final_json_output)
        
    print(f"Arquivo '{output_file}' unificado gerado com sucesso, contendo {len(unified_data)} publicações.")
